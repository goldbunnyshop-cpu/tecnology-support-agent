# agent/reports.py — Generador de reporte Excel semanal de leads
# Generado por AgentKit

import os
import logging
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from anthropic import AsyncAnthropic
from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger("agentkit")
client = AsyncAnthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

ZONA = ZoneInfo("America/Mexico_City")
DIRECTORIO_REPORTES = "reports"

# Colores por estado del funnel
COLORES = {
    "activo":         "D6E4FF",  # azul claro
    "en_seguimiento": "FFF3CD",  # amarillo
    "perdido":        "FFD6D6",  # rojo claro
    "convertido":     "D6F5D6",  # verde claro
}

ETIQUETAS = {
    "activo":         "Activo",
    "en_seguimiento": "En seguimiento",
    "perdido":        "Perdido",
    "convertido":     "Convertido",
}


async def extraer_info_conversacion(historial: list[dict]) -> dict:
    """
    Usa Claude para extraer nombre, dispositivo y resumen del problema
    de una conversación, en una sola llamada rápida.
    """
    if not historial:
        return {"nombre": "Desconocido", "dispositivo": "No especificado", "resumen": "Sin detalles"}

    fragmento = "\n".join(
        f"{'Cliente' if m['role'] == 'user' else 'Agente'}: {m['content']}"
        for m in historial[-10:]
    )

    prompt = f"""Analiza esta conversación de WhatsApp de un taller de reparación y extrae:
1. Nombre del cliente (si lo mencionó, si no: "Desconocido")
2. Dispositivo (ej: "iPhone 13", "Samsung S22", "PS5", "Laptop HP", si no: "No especificado")
3. Resumen del problema en máximo 8 palabras

Conversación:
{fragmento}

Responde SOLO en este formato exacto (sin explicaciones):
Nombre: [valor]
Dispositivo: [valor]
Problema: [valor]"""

    try:
        response = await client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=80,
            messages=[{"role": "user", "content": prompt}]
        )
        texto = response.content[0].text.strip()
        resultado = {"nombre": "Desconocido", "dispositivo": "No especificado", "resumen": "Sin detalles"}
        for linea in texto.splitlines():
            if linea.startswith("Nombre:"):
                resultado["nombre"] = linea.replace("Nombre:", "").strip()
            elif linea.startswith("Dispositivo:"):
                resultado["dispositivo"] = linea.replace("Dispositivo:", "").strip()
            elif linea.startswith("Problema:"):
                resultado["resumen"] = linea.replace("Problema:", "").strip()
        return resultado
    except Exception as e:
        logger.error(f"Error extrayendo info de conversación: {e}")
        return {"nombre": "Desconocido", "dispositivo": "No especificado", "resumen": "Sin detalles"}


def _estilo_encabezado(celda, color_hex: str = "1F4E79"):
    celda.font = Font(bold=True, color="FFFFFF", size=11)
    celda.fill = PatternFill("solid", fgColor=color_hex)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Side(style="thin", color="CCCCCC")
    celda.border = Border(left=borde, right=borde, top=borde, bottom=borde)


def _estilo_celda(celda, color_fondo: str = "FFFFFF"):
    celda.fill = PatternFill("solid", fgColor=color_fondo)
    celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde = Side(style="thin", color="DDDDDD")
    celda.border = Border(left=borde, right=borde, top=borde, bottom=borde)


def _ajustar_columnas(hoja, anchos: list[int]):
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho


def _agregar_hoja_leads(wb, nombre_hoja: str, filas: list[dict], color_estado: str, mostrar_detalle_anuncio: bool = False):
    """Crea una hoja con los leads de un estado específico."""
    ws = wb.create_sheet(title=nombre_hoja)

    encabezados = [
        "Teléfono", "Nombre", "Dispositivo", "Problema",
        "Fuente", "Último mensaje", "Días sin respuesta", "Seguimientos enviados", "Registrado"
    ]
    anchos = [18, 20, 20, 30, 18, 22, 20, 22, 22]

    if mostrar_detalle_anuncio:
        encabezados.append("Anuncio / Detalle")
        anchos.append(40)

    ws.row_dimensions[1].height = 30
    for col, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        _estilo_encabezado(celda)

    ahora = datetime.now(ZONA)

    for fila_num, datos in enumerate(filas, start=2):
        ultimo = datos.get("ultimo_mensaje")
        if ultimo:
            if ultimo.tzinfo is None:
                ultimo = ultimo.replace(tzinfo=timezone.utc)
            dias_sin = (ahora - ultimo.astimezone(ZONA)).days
            ultimo_str = ultimo.astimezone(ZONA).strftime("%d/%m/%Y %H:%M")
        else:
            dias_sin = "-"
            ultimo_str = "-"

        registrado = datos.get("created_at")
        registrado_str = (
            registrado.astimezone(ZONA).strftime("%d/%m/%Y")
            if registrado and registrado.tzinfo
            else (registrado.strftime("%d/%m/%Y") if registrado else "-")
        )

        fuente_raw = datos.get("fuente", "desconocido") or "desconocido"
        fuente_label = {
            "facebook_ad": "Facebook Ad",
            "instagram_ad": "Instagram Ad",
            "organico": "Orgánico",
            "referido": "Referido",
            "desconocido": "—",
        }.get(fuente_raw, fuente_raw)

        valores = [
            datos.get("telefono", "-"),
            datos.get("nombre", "Desconocido"),
            datos.get("dispositivo", "No especificado"),
            datos.get("resumen", "Sin detalles"),
            fuente_label,
            ultimo_str,
            dias_sin,
            datos.get("seguimientos_enviados", 0),
            registrado_str,
        ]

        if mostrar_detalle_anuncio:
            valores.append(datos.get("fuente_detalle", "") or "")

        ws.row_dimensions[fila_num].height = 20
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_num, column=col, value=valor)
            _estilo_celda(celda, color_estado)

    _ajustar_columnas(ws, anchos)
    return ws


def _agregar_hoja_resumen(wb, conteos: dict, semana: str):
    """Crea la hoja de resumen con totales y métricas clave."""
    ws = wb.create_sheet(title="Resumen", index=0)

    ws.merge_cells("A1:D1")
    titulo = ws["A1"]
    titulo.value = f"Tecnology Support — Reporte Semanal de Leads ({semana})"
    titulo.font = Font(bold=True, color="FFFFFF", size=14)
    titulo.fill = PatternFill("solid", fgColor="1F4E79")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.append([])

    encabezados = ["Estado", "Cantidad", "% del total", "Acción sugerida"]
    for col, h in enumerate(encabezados, start=1):
        celda = ws.cell(row=3, column=col, value=h)
        _estilo_encabezado(celda, "2E75B6")
    ws.row_dimensions[3].height = 25

    total = conteos.get("total", 1) or 1
    acciones = {
        "Activos":          "Monitorear — pueden convertir esta semana",
        "En seguimiento":   "Revisar manualmente — necesitan atencion personal",
        "Perdidos":         "Reactivar con campana puntual si aplica",
        "Convertidos":      "Solicitar resena o referido",
    }
    orden = [
        ("Activos",          conteos.get("activo", 0),         COLORES["activo"]),
        ("En seguimiento",   conteos.get("en_seguimiento", 0), COLORES["en_seguimiento"]),
        ("Perdidos",         conteos.get("perdido", 0),         COLORES["perdido"]),
        ("Convertidos",      conteos.get("convertido", 0),      COLORES["convertido"]),
    ]

    for fila_num, (etiqueta, cantidad, color) in enumerate(orden, start=4):
        porcentaje = f"{cantidad / total * 100:.1f}%"
        valores = [etiqueta, cantidad, porcentaje, acciones[etiqueta]]
        ws.row_dimensions[fila_num].height = 20
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_num, column=col, value=valor)
            _estilo_celda(celda, color)

    # Fila de total
    fila_total = len(orden) + 4
    ws.row_dimensions[fila_total].height = 22
    for col, valor in enumerate(["TOTAL", total, "100%", ""], start=1):
        celda = ws.cell(row=fila_total, column=col, value=valor)
        celda.font = Font(bold=True, size=11)
        celda.fill = PatternFill("solid", fgColor="E2EFDA")
        celda.alignment = Alignment(horizontal="left", vertical="center")

    _ajustar_columnas(ws, [22, 12, 14, 45])

    # Nota al pie
    ws.cell(row=fila_total + 2, column=1,
            value=f"Generado automaticamente el {datetime.now(ZONA).strftime('%d/%m/%Y a las %H:%M')} hora Ciudad de Mexico")
    ws.cell(row=fila_total + 2, column=1).font = Font(italic=True, color="888888", size=9)

    return ws


async def generar_reporte_excel() -> str:
    """
    Genera el archivo Excel con el reporte semanal de leads.
    Retorna la ruta del archivo generado.
    """
    from agent.leads import obtener_todos_los_leads, obtener_resumen_leads
    from agent.memory import obtener_historial

    logger.info("Generando reporte semanal de leads...")

    leads = await obtener_todos_los_leads()
    conteos = await obtener_resumen_leads()

    # Agrupar leads por estado y enriquecer con info de conversación
    grupos: dict[str, list[dict]] = {
        "activo": [], "en_seguimiento": [], "perdido": [], "convertido": []
    }

    facebook_ads = []
    FECHA_INICIO_ADS = datetime(2025, 3, 19, tzinfo=timezone.utc)

    for lead in leads:
        historial = await obtener_historial(lead.telefono, limite=10)
        info = await extraer_info_conversacion(historial)
        fuente = getattr(lead, "fuente", "desconocido") or "desconocido"
        fila = {
            "telefono": lead.telefono,
            "nombre": info["nombre"],
            "dispositivo": info["dispositivo"],
            "resumen": info["resumen"],
            "fuente": fuente,
            "fuente_detalle": getattr(lead, "fuente_detalle", "") or "",
            "ultimo_mensaje": lead.ultimo_mensaje,
            "seguimientos_enviados": lead.seguimientos_enviados,
            "created_at": lead.created_at,
        }
        grupos.get(lead.estado, grupos["activo"]).append(fila)
        if fuente in ("facebook_ad", "instagram_ad"):
            # Solo incluir leads desde el 19 de marzo
            created = lead.created_at
            if created:
                if created.tzinfo is None:
                    created = created.replace(tzinfo=timezone.utc)
                if created >= FECHA_INICIO_ADS:
                    facebook_ads.append(fila)
            else:
                facebook_ads.append(fila)

    # Construir Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    semana = datetime.now(ZONA).strftime("Semana del %d/%m/%Y")
    _agregar_hoja_resumen(wb, conteos, semana)
    # Hoja exclusiva de Facebook/Instagram Ads (color naranja)
    if facebook_ads:
        _agregar_hoja_leads(wb, "Facebook & Instagram Ads", facebook_ads, "FFE0C0", mostrar_detalle_anuncio=True)
    _agregar_hoja_leads(wb, "Activos",          grupos["activo"],         COLORES["activo"])
    _agregar_hoja_leads(wb, "En Seguimiento",    grupos["en_seguimiento"], COLORES["en_seguimiento"])
    _agregar_hoja_leads(wb, "Perdidos",          grupos["perdido"],        COLORES["perdido"])
    _agregar_hoja_leads(wb, "Convertidos",       grupos["convertido"],     COLORES["convertido"])

    # Guardar archivo
    os.makedirs(DIRECTORIO_REPORTES, exist_ok=True)
    nombre_archivo = f"reporte_leads_{datetime.now(ZONA).strftime('%Y-%m-%d')}.xlsx"
    ruta = os.path.join(DIRECTORIO_REPORTES, nombre_archivo)
    wb.save(ruta)

    total = conteos.get("total", 0)
    logger.info(f"Reporte generado: {ruta} ({total} leads, {len(leads)} procesados)")
    return ruta
